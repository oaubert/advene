#
# Advene: Annotate Digital Videos, Exchange on the NEt
# Copyright (C) 2024 Olivier Aubert <contact@olivieraubert.net>
#
# Advene is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# Advene is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Advene; if not, write to the Free Software
# Foundation, Inc., 51 Franklin St, Fifth Floor, Boston, MA  02110-1301  USA
#
"""Corpus XLSX export export filter.

This filter exports corpus data (cross-package) as XLSX files.
"""

name = "Corpus Xlsx exporter"

import logging
logger = logging.getLogger(__name__)

from gettext import gettext as _

import os

import advene.core.config as config
from advene.util.exporter import GenericExporter

import advene.util.helper as helper

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None

def register(controller=None):
    if openpyxl:
        controller.register_exporter(CorpusXlsxExporter)
    return True

class CorpusXlsxExporter(GenericExporter):
    """Corpus Xlsx Exporter
    """
    name = _("Corpus Xlsx exporter")
    extension = 'xlsx'
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def __init__(self, controller=None, source=None, callback=None):
        super().__init__(controller, source, callback)
        self.sheet_global = True
        self.sheet_relationtypes_summary = False
        self.sheet_by_package = False
        self.sheet_by_annotationtype = True
        self.sheet_by_relationtype = False
        package_count = len(controller.global_package.package_list)
        self.all = controller.global_package
        self.optionparser.description = _(f"Export {package_count} packages as Xlsx file.")
        self.optionparser.add_option("-r", "--relation-types",
                                     action="store_true", dest="sheet_by_relationtype", default=self.sheet_by_relationtype,
                                     help=_("Create 1 tab per relation-type"))
        self.optionparser.add_option("-a", "--annotation-types",
                                     action="store_true", dest="sheet_by_annotationtype", default=self.sheet_by_annotationtype,
                                     help=_("Create 1 tab per annotation-type"))
        self.optionparser.add_option("-p", "--packages",
                                     action="store_true", dest="sheet_by_package", default=self.sheet_by_package,
                                     help=_("Create 1 tab per package"))
        self.optionparser.add_option("-s", "--relation-types-summary",
                                     action="store_true", dest="sheet_relationtypes_summary", default=self.sheet_relationtypes_summary,
                                     help=_("Create a tab with relation types summaries"))
        self.optionparser.add_option("-g", "--global",
                                     action="store_true", dest="sheet_global", default=self.sheet_global,
                                     help=_("Create 1 global sheet with all annotations"))

    @classmethod
    def is_valid_for(cls, expr):
        return expr == 'package'

    def fit_column_width(self, sheet):
        # Iterate over all columns and adjust their widths
        for column in sheet.columns:
            column_letter = column[0].column_letter
            max_length = max(len(str(cell.value))
                             for cell in column)
            sheet.column_dimensions[column_letter].width = min(50,
                                                               max_length + 2)

    def export(self, filename=None):
        # Export a corpus as a Xlsx file

        book = openpyxl.Workbook()

        # Remove default worksheet
        book.remove(book.worksheets[0])

        # Add summary worksheet
        sheet = book.create_sheet('Summary')
        header = [
            "alias",
            "title",
            "media",
            "annotation count",
            "annotation type count",
            "relation type count",
            "media duration"
        ]
        sheet.append(header)
        for alias, p in self.controller.packages.items():
            if alias == 'advene':
                continue
            # Store alias as p attribute, for reference in next steps
            p.alias = alias
            media_uri = os.path.basename(p.getMetaData(config.data.namespace, "media_uri") or "unknown")
            sheet.append([
                alias,
                p.title,
                media_uri,
                len(p.annotations),
                len(p.annotationTypes),
                len(p.relationTypes),
                p.getMetaData(config.data.namespace, "duration") or ""
            ])
            for row in sheet.iter_rows(min_row=1, max_row=1, max_col=1 + len(header)):
                for cell in row:
                    cell.font = Font(bold=True)

        self.fit_column_width(sheet)

        def save_annotations(ws, annotations):
            header = [ "text", "begin", "begin_ms", "end", "end_ms", "type_id", "type_title", "id", "package" ]
            ws.append(header)
            for row in sheet.iter_rows(min_row=1, max_row=1, max_col=1 + len(header)):
                for cell in row:
                    cell.font = Font(bold=True)
            for a in annotations:
                ws.append([
                    self.controller.get_title(a),
                    helper.format_time_reference(a.fragment.begin),
                    a.fragment.begin,
                    helper.format_time_reference(a.fragment.end),
                    a.fragment.end,
                    a.type.id,
                    self.controller.get_title(a.type),
                    a.id,
                    a.ownerPackage.alias
                ])
            self.fit_column_width(sheet)

        # Add global sheet
        if self.sheet_global:
            sheet = book.create_sheet('Global')
            save_annotations(sheet, sorted(self.controller.global_package.annotations))

        # Add a relations sheets
        # A single sheet - For each package, 1 row per relation, 1 column by annotation type
        if self.sheet_relationtypes_summary:
            sheet = book.create_sheet("Relation types")
            at_titles = set()
            # First pass to determine used annotation types
            for p in self.controller.global_package.package_list:
                for rt in p.relationTypes:
                    # Update the available annotation types titles
                    at_titles.update(a.type.title for a in rt.annotations)
            sheet.append([ 'Package', 'Package title', 'Relation type title', 'Summary', *list(at_titles) ])
            for alias, p in self.controller.global_package.package_dict.items():
                for rt in p.relationTypes:
                    # Build the content for each annotation type title (concatenate annotations)
                    annotations = rt.annotations
                    if annotations:
                        contents = dict((at_title, "\n".join(self.controller.get_title(a)
                                                             for a in annotations
                                                             if a.type.title == at_title
                                                             ))
                                        for at_title in at_titles)
                        summary = " - ".join(self.controller.get_title(a)
                                             for a in annotations
                                             # FIXME: not very generic - find a better way to generate summary
                                             if a.type.title != 'Verbalisations')
                        sheet.append([ alias,
                                       p.title,
                                       self.controller.get_title(rt),
                                       summary,
                                       *(contents.get(at_title, "") for at_title in at_titles)])

        # Create 1 sheet per package
        if self.sheet_by_package:
            for alias, p in self.controller.packages.items():
                if alias == 'advene':
                    continue
                sheet = book.create_sheet(alias)

                media_uri = p.getMetaData(config.data.namespace, "media_uri") or "unknown"

                save_annotations(sheet, p.annotations)

        # Create 1 sheet per annotation type (label)
        if self.sheet_by_annotationtype:
            at_titles = set(at.title for at in self.controller.global_package.annotationTypes)
            for at_title in sorted(at_titles):
                sheet = book.create_sheet(at_title)
                save_annotations(sheet, (a
                                         for alias, p in self.controller.packages.items()
                                         if alias != 'advene'
                                         for a in sorted(p.annotations)
                                         if a.type.title == at_title))

        # Create 1 sheet per relation type (label)
        if self.sheet_by_relationtype:
            rt_titles = set(rt.title for rt in self.controller.global_package.relationTypes)
            for rt_title in sorted(rt_titles):
                sheet = book.create_sheet(rt_title)
                save_annotations(sheet, (a
                                         for alias, p in self.controller.packages.items()
                                         if alias != 'advene'
                                         for a in sorted(p.annotations)
                                         if rt_title in (r.type.title for r in a.relations)))

        book.save(filename)
        return filename
